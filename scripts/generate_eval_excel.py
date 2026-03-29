"""Generate Excel file from maritime evaluation dataset (300 questions).

Usage:
    PYTHONPATH=. python3 scripts/generate_eval_excel.py
"""

import sys
from pathlib import Path

# Ensure project root is on path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from domains.maritime.evaluation.dataset import EvalDataset, Difficulty

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
COLOUR_ALL = "BBDEFB"       # light blue  (전체)
COLOUR_EASY = "E8F5E9"      # light green
COLOUR_MEDIUM = "FFF3E0"    # light orange
COLOUR_HARD = "FFEBEE"      # light red
COLOUR_SUMMARY = "F3E5F5"   # light purple (요약)

# Map difficulty → hex colour (no '#')
DIFF_COLOUR = {
    "EASY": COLOUR_EASY,
    "MEDIUM": COLOUR_MEDIUM,
    "HARD": COLOUR_HARD,
}

COLUMNS = [
    "#",
    "질문 (Question)",
    "추론유형 (ReasoningType)",
    "난이도 (Difficulty)",
    "예상 레이블 (Expected Labels)",
    "Ground Truth Cypher",
    "설명 (Description)",
]

# Approximate column widths (characters)
COL_WIDTHS = [6, 45, 18, 14, 35, 70, 35]


def make_thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_header(ws, colour_hex: str) -> None:
    fill = PatternFill("solid", fgColor=colour_hex)
    bold_font = Font(bold=True, size=10)
    border = make_thin_border()

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = bold_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28


def write_question_rows(ws, questions, header_colour_hex: str) -> None:
    border = make_thin_border()
    normal_font = Font(size=10)

    for row_num, (idx, q) in enumerate(enumerate(questions, start=1), start=2):
        row_colour = header_colour_hex  # use same tint for odd rows
        # Alternate row background: slightly lighter for even rows
        fill = PatternFill("solid", fgColor=row_colour) if row_num % 2 == 0 else None

        values = [
            idx,
            q.question,
            q.reasoning_type.value,
            q.difficulty.value,
            ", ".join(q.expected_labels),
            q.ground_truth_cypher,
            q.description,
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col_idx in (1, 3, 4) else "left",
            )
            if fill:
                cell.fill = fill


def set_column_widths(ws) -> None:
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def build_question_sheet(wb: Workbook, title: str, colour_hex: str, questions) -> None:
    ws = wb.create_sheet(title=title)
    write_header(ws, colour_hex)
    write_question_rows(ws, questions, colour_hex)
    set_column_widths(ws)
    freeze_header(ws)


def build_summary_sheet(wb: Workbook, dataset: EvalDataset) -> None:
    ws = wb.create_sheet(title="요약 (Summary)")
    header_fill = PatternFill("solid", fgColor=COLOUR_SUMMARY)
    bold = Font(bold=True, size=10)
    normal = Font(size=10)
    border = make_thin_border()

    def hdr_cell(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = header_fill
        c.font = bold
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    def data_cell(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = normal
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Section 1: 난이도별 통계 ---
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "난이도별 통계 (By Difficulty)"
    title_cell.font = Font(bold=True, size=11)
    title_cell.fill = PatternFill("solid", fgColor="E1BEE7")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border

    hdr_cell(2, 1, "난이도")
    hdr_cell(2, 2, "개수")
    hdr_cell(2, 3, "비율")

    total = len(dataset.questions)
    diff_rows = {
        "EASY": (Difficulty.EASY, COLOUR_EASY),
        "MEDIUM": (Difficulty.MEDIUM, COLOUR_MEDIUM),
        "HARD": (Difficulty.HARD, COLOUR_HARD),
    }

    for r, (diff_name, (diff_enum, colour)) in enumerate(diff_rows.items(), start=3):
        count = len(dataset.get_by_difficulty(diff_enum))
        pct = f"{count / total * 100:.1f}%"

        c_name = ws.cell(row=r, column=1, value=diff_name)
        c_name.fill = PatternFill("solid", fgColor=colour)
        c_name.font = Font(bold=True, size=10)
        c_name.border = border
        c_name.alignment = Alignment(horizontal="center", vertical="center")

        data_cell(r, 2, count)
        data_cell(r, 3, pct)

    # Total row
    total_cell = ws.cell(row=6, column=1, value="합계")
    total_cell.font = Font(bold=True, size=10)
    total_cell.border = border
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    data_cell(6, 2, total)
    data_cell(6, 3, "100.0%")

    # --- Section 2: 추론유형별 통계 ---
    ws.merge_cells("E1:G1")
    title2 = ws["E1"]
    title2.value = "추론유형별 통계 (By Reasoning Type)"
    title2.font = Font(bold=True, size=11)
    title2.fill = PatternFill("solid", fgColor="E1BEE7")
    title2.alignment = Alignment(horizontal="center", vertical="center")
    title2.border = border

    hdr_cell(2, 5, "추론유형")
    hdr_cell(2, 6, "개수")
    hdr_cell(2, 7, "비율")

    from domains.maritime.evaluation.dataset import ReasoningType

    reasoning_colours = {
        "DIRECT": "E3F2FD",
        "BRIDGE": "E8F5E9",
        "COMPARISON": "FFF3E0",
        "INTERSECTION": "FFEBEE",
        "COMPOSITION": "F3E5F5",
    }

    for r, rt in enumerate(ReasoningType, start=3):
        count = len(dataset.get_by_reasoning_type(rt))
        pct = f"{count / total * 100:.1f}%"
        colour = reasoning_colours.get(rt.value, "FFFFFF")

        c_rt = ws.cell(row=r, column=5, value=rt.value)
        c_rt.fill = PatternFill("solid", fgColor=colour)
        c_rt.font = Font(bold=True, size=10)
        c_rt.border = border
        c_rt.alignment = Alignment(horizontal="center", vertical="center")

        data_cell(r, 6, count)
        data_cell(r, 7, pct)

    # Total row for reasoning
    total_rt_row = 3 + len(list(ReasoningType))
    t2 = ws.cell(row=total_rt_row, column=5, value="합계")
    t2.font = Font(bold=True, size=10)
    t2.border = border
    t2.alignment = Alignment(horizontal="center", vertical="center")
    data_cell(total_rt_row, 6, total)
    data_cell(total_rt_row, 7, "100.0%")

    # --- Section 3: Cross-table (난이도 × 추론유형) ---
    cross_start_row = 9

    ws.merge_cells(f"A{cross_start_row}:G{cross_start_row}")
    cross_title = ws.cell(row=cross_start_row, column=1, value="교차 통계 (Difficulty × ReasoningType)")
    cross_title.font = Font(bold=True, size=11)
    cross_title.fill = PatternFill("solid", fgColor="E1BEE7")
    cross_title.alignment = Alignment(horizontal="center", vertical="center")
    cross_title.border = border

    # Header row
    hdr_cell(cross_start_row + 1, 1, "난이도 \\ 추론유형")
    for c_idx, rt in enumerate(ReasoningType, start=2):
        hdr_cell(cross_start_row + 1, c_idx, rt.value)
    hdr_cell(cross_start_row + 1, len(list(ReasoningType)) + 2, "합계")

    # Data rows
    for r_idx, (diff_name, (diff_enum, colour)) in enumerate(diff_rows.items(), start=cross_start_row + 2):
        c_name = ws.cell(row=r_idx, column=1, value=diff_name)
        c_name.fill = PatternFill("solid", fgColor=colour)
        c_name.font = Font(bold=True, size=10)
        c_name.border = border
        c_name.alignment = Alignment(horizontal="center", vertical="center")

        row_total = 0
        for c_idx, rt in enumerate(ReasoningType, start=2):
            count = sum(
                1 for q in dataset.questions
                if q.difficulty == diff_enum and q.reasoning_type == rt
            )
            row_total += count
            data_cell(r_idx, c_idx, count)

        data_cell(r_idx, len(list(ReasoningType)) + 2, row_total)

    # Column totals
    total_cross_row = cross_start_row + 2 + len(diff_rows)
    t3 = ws.cell(row=total_cross_row, column=1, value="합계")
    t3.font = Font(bold=True, size=10)
    t3.border = border
    t3.alignment = Alignment(horizontal="center", vertical="center")

    grand = 0
    for c_idx, rt in enumerate(ReasoningType, start=2):
        count = len(dataset.get_by_reasoning_type(rt))
        grand += count
        data_cell(total_cross_row, c_idx, count)
    data_cell(total_cross_row, len(list(ReasoningType)) + 2, grand)

    # Column widths for summary sheet
    summary_widths = [28, 14, 14, 14, 22, 10, 10]
    for col_idx, w in enumerate(summary_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Row heights
    ws.row_dimensions[1].height = 24
    for r in range(2, total_cross_row + 1):
        ws.row_dimensions[r].height = 20


def main() -> None:
    output_path = PROJECT_ROOT / "docs" / "evaluation_questions_300.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    dataset = EvalDataset.builtin()
    print(dataset.summary())
    print()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. 전체 (All)
    build_question_sheet(wb, "전체 (All)", COLOUR_ALL, dataset.questions)

    # 2-4. Per difficulty
    for diff, colour in [
        (Difficulty.EASY, COLOUR_EASY),
        (Difficulty.MEDIUM, COLOUR_MEDIUM),
        (Difficulty.HARD, COLOUR_HARD),
    ]:
        questions = dataset.get_by_difficulty(diff)
        build_question_sheet(wb, diff.value, colour, questions)

    # 5. Summary
    build_summary_sheet(wb, dataset)

    wb.save(str(output_path))
    print(f"Saved: {output_path}")
    print(f"Total questions written: {len(dataset.questions)}")


if __name__ == "__main__":
    main()
